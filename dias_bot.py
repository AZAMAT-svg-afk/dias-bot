#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║        DIAS TARGETPRO — BOT + API SERVER v5.0                       ║
║        ✅ Flask HTTP сервер  ✅ Telegram бот  ✅ Google Sheets       ║
║        🔒 Все секреты — только через переменные окружения           ║
╚══════════════════════════════════════════════════════════════════════╝

Переменные окружения (обязательные):
    BOT_TOKEN        — токен Telegram-бота
    ADMIN_ID         — Telegram chat_id администратора
    SECRET           — секрет для защиты /api/submit
    APPS_SCRIPT_URL  — URL Google Apps Script Web App
    ADMIN_PASSWORD_HASH — SHA-256 хэш пароля для admin-панели

Переменные окружения (опциональные):
    WA_NUMBER        — номер WhatsApp для кнопки (без +, формат 77001234567)
    TG_BOT_TOKEN_PUBLIC — токен бота для уведомлений с сайта (может совпадать с BOT_TOKEN)
    TG_CHAT_ID_PUBLIC   — chat_id куда слать уведомления с сайта
    PORT             — порт Flask сервера (по умолчанию 5000)
    SHEET_ID         — ID Google Таблицы (для чтения статистики)
    SHEET_QUIZ_GID   — GID листа с квизами

Запуск локально:
    pip install flask pyTelegramBotAPI openpyxl python-docx requests
    export BOT_TOKEN=...
    export ADMIN_ID=...
    export SECRET=...
    export APPS_SCRIPT_URL=...
    export ADMIN_PASSWORD_HASH=...
    python dias_bot.py

Render.com:
    Start Command: python dias_bot.py
    Port: 5000
    Environment Variables: задать все обязательные переменные в Dashboard
"""

import telebot
import json
import os
import io
import threading
import urllib.request
import requests
from datetime import datetime, timedelta
from collections import defaultdict
from flask import Flask, request, jsonify

# ╔═══════════════════════════════════════╗
# ║           НАСТРОЙКИ — ТОЛЬКО ИЗ ENV  ║
# ╚═══════════════════════════════════════╝

def _require_env(name):
    """Получает переменную окружения; падает при старте если не задана."""
    val = os.environ.get(name, '').strip()
    if not val:
        raise RuntimeError(
            f"\n❌  Переменная окружения '{name}' не задана!\n"
            f"    Задайте её перед запуском: export {name}=...\n"
            f"    На Render.com: Dashboard → Environment Variables"
        )
    return val

BOT_TOKEN          = _require_env('BOT_TOKEN')
ADMIN_ID           = _require_env('ADMIN_ID')
SECRET             = _require_env('SECRET')
APPS_SCRIPT_URL    = _require_env('APPS_SCRIPT_URL')
ADMIN_PASSWORD_HASH = _require_env('ADMIN_PASSWORD_HASH')

# Опциональные
WA_NUMBER           = os.environ.get('WA_NUMBER', '').strip()
TG_BOT_TOKEN_PUBLIC = os.environ.get('TG_BOT_TOKEN_PUBLIC', BOT_TOKEN).strip()
TG_CHAT_ID_PUBLIC   = os.environ.get('TG_CHAT_ID_PUBLIC', ADMIN_ID).strip()
PORT                = int(os.environ.get('PORT', 5000))
SHEET_ID            = os.environ.get('SHEET_ID', '').strip()
SHEET_QUIZ_GID      = os.environ.get('SHEET_QUIZ_GID', '').strip()

DB_FILE = 'quiz_db.json'

# ════════════════════════════════════════
app = Flask(__name__)
bot = telebot.TeleBot(BOT_TOKEN)

# ════════════════════════════════════════
#  ЛОКАЛЬНАЯ БАЗА ДАННЫХ (JSON файл)
# ════════════════════════════════════════

def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'quizzes': []}

def save_db(db):
    with open(DB_FILE, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def add_quiz_entry(data):
    db = load_db()
    entry = {
        'name':          data.get('name', '—'),
        'phone':         data.get('phone', '—'),
        'business':      data.get('business', '—'),
        'city':          data.get('city', '—'),
        'budget':        data.get('budget', '—'),
        'avg_check':     data.get('avgCheck', '—'),
        'client_source': data.get('clientSource', '—'),
        'main_problem':  data.get('mainProblem', '—'),
        'has_sales':     data.get('hasSales', '—'),
        'goal':          data.get('goal', '—'),
        'lang':          data.get('lang', 'ru'),
        'utm':           json.dumps(data.get('utm', {}), ensure_ascii=False),
        'time':          _now_str(),
        'timestamp':     datetime.now().isoformat(),
        'status':        'pending',
    }
    db['quizzes'].append(entry)
    save_db(db)
    return entry

def _clean_phone(s):
    return str(s).replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')

def get_status_by_phone(phone):
    db = load_db()
    search = _clean_phone(phone)
    for q in reversed(db.get('quizzes', [])):
        if _clean_phone(q.get('phone', '')) == search or q.get('phone', '') == phone:
            return q.get('status', 'pending')
    return 'pending'

def set_status_by_phone(phone, status):
    db = load_db()
    search = _clean_phone(phone)
    changed = False
    for q in db.get('quizzes', []):
        if _clean_phone(q.get('phone', '')) == search or q.get('phone', '') == phone:
            q['status'] = status
            changed = True
    if changed:
        save_db(db)
    return changed

def delete_by_phone(phone):
    db = load_db()
    search = _clean_phone(phone)
    before = len(db.get('quizzes', []))
    db['quizzes'] = [
        q for q in db.get('quizzes', [])
        if _clean_phone(q.get('phone', '')) != search and q.get('phone', '') != phone
    ]
    if len(db['quizzes']) < before:
        save_db(db)
        return True
    return False

def _now_str():
    return datetime.now().strftime('%d.%m.%Y %H:%M')

# ════════════════════════════════════════
#  ПРОКСИ К GOOGLE APPS SCRIPT
#  (URL Apps Script не покидает сервер)
# ════════════════════════════════════════

def forward_to_sheets(data):
    """Пересылает данные квиза в Google Apps Script → Google Sheets."""
    try:
        payload = dict(data)
        payload['secret'] = SECRET
        payload['type'] = 'quiz'
        res = requests.post(
            APPS_SCRIPT_URL,
            data=json.dumps(payload),
            headers={'Content-Type': 'text/plain'},
            timeout=15
        )
        result = res.json()
        if result.get('ok'):
            print(f"✅ Sheets: строка добавлена (row {result.get('row', '?')})")
            return True
        else:
            print(f"⚠️  Sheets вернул: {result}")
            return False
    except Exception as e:
        print(f"❌ Ошибка пересылки в Sheets: {e}")
        return False

def check_status_in_sheets(phone):
    """Проверяет статус заявки через Apps Script doGet."""
    try:
        url = f"{APPS_SCRIPT_URL}?phone={requests.utils.quote(str(phone))}"
        res = requests.get(url, timeout=10)
        data = res.json()
        return data.get('status', 'pending')
    except Exception as e:
        print(f"❌ Ошибка проверки статуса: {e}")
        return None

def proxy_to_sheets(payload):
    """
    Проксирует действия admin-панели (list / approve / reject / delete)
    в Apps Script, добавляя secret на сервере.
    """
    payload['secret'] = SECRET
    try:
        res = requests.post(
            APPS_SCRIPT_URL,
            data=json.dumps(payload),
            headers={'Content-Type': 'text/plain'},
            timeout=15
        )
        return res.json()
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def proxy_get_sheets(params):
    """
    Проксирует GET-запросы к Apps Script (для списка заявок).
    """
    params['secret'] = SECRET
    try:
        res = requests.get(APPS_SCRIPT_URL, params=params, timeout=15)
        return res.json()
    except Exception as e:
        return {'ok': False, 'error': str(e)}

# ════════════════════════════════════════
#  FLASK — HTTP API
# ════════════════════════════════════════

@app.route('/', methods=['GET'])
def index():
    return jsonify({'status': 'ok', 'service': 'Dias TargetPro API v5.0'})

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'ok': True})

# ─── Конфиг для фронтенда (БЕЗ секретов) ───
@app.route('/api/config', methods=['GET'])
def api_config():
    """
    Отдаёт фронтенду только то, что нужно для работы.
    Никаких токенов, URL Apps Script и секретов здесь нет.
    Всё взаимодействие с внешними сервисами идёт через бэкенд.
    """
    return _cors_response(jsonify({
        'backend_url':     '/api/submit',
        'status_url':      '/api/status',
        'revoke_url':      '/api/revoke',
        'wa_number':       WA_NUMBER,
        # tg_bot_token и apps_script_url фронтенду НЕ нужны —
        # все уведомления теперь отправляет сервер
    }))

# ─── Проксирование запросов admin-панели к Apps Script ───
@app.route('/api/admin/leads', methods=['GET', 'OPTIONS'])
def api_admin_leads():
    """Список заявок для admin-панели — прокси к Apps Script."""
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    if not _check_admin_session():
        return _cors_response(jsonify({'ok': False, 'error': 'unauthorized'}), 403)

    params = {'action': 'list'}
    result = proxy_get_sheets(params)
    return _cors_response(jsonify(result))

@app.route('/api/admin/action', methods=['POST', 'OPTIONS'])
def api_admin_action():
    """Действия admin-панели (approve / reject / delete) — прокси к Apps Script."""
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    if not _check_admin_session():
        return _cors_response(jsonify({'ok': False, 'error': 'unauthorized'}), 403)

    data = request.get_json(force=True, silent=True) or {}
    action_type = data.get('type', '')

    if action_type not in ('approve', 'reject', 'delete'):
        return _cors_response(jsonify({'ok': False, 'error': 'invalid_action'}), 400)

    payload = {
        'type':  action_type,
        'row':   data.get('row'),
        'phone': data.get('phone', ''),
    }
    result = proxy_to_sheets(payload)
    return _cors_response(jsonify(result))

@app.route('/api/admin/login', methods=['POST', 'OPTIONS'])
def api_admin_login():
    """
    Верификация пароля admin-панели на сервере.
    Фронтенд присылает SHA-256 хэш пароля — сравниваем с ADMIN_PASSWORD_HASH из env.
    """
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    data = request.get_json(force=True, silent=True) or {}
    pwd_hash = data.get('hash', '').lower().strip()

    if not pwd_hash or pwd_hash != ADMIN_PASSWORD_HASH.lower().strip():
        return _cors_response(jsonify({'ok': False, 'error': 'wrong_password'}), 403)

    # Простой session token — для production используйте Flask-Login или JWT
    import hashlib, time
    session_token = hashlib.sha256(f"{SECRET}{time.time()}".encode()).hexdigest()
    # В production храните токен в Redis/DB; здесь для простоты — в памяти
    _admin_sessions.add(session_token)
    return _cors_response(jsonify({'ok': True, 'token': session_token}))

@app.route('/api/admin/logout', methods=['POST', 'OPTIONS'])
def api_admin_logout():
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))
    token = request.headers.get('X-Admin-Token', '')
    _admin_sessions.discard(token)
    return _cors_response(jsonify({'ok': True}))

# Хранилище сессий в памяти (сбрасывается при перезапуске)
_admin_sessions = set()

def _check_admin_session():
    token = request.headers.get('X-Admin-Token', '')
    return token in _admin_sessions

# ─── Приём заявки с сайта ───
@app.route('/api/submit', methods=['POST', 'OPTIONS'])
def api_submit():
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    try:
        data = request.get_json(force=True, silent=True) or {}

        if data.get('secret') != SECRET:
            return _cors_response(jsonify({'ok': False, 'error': 'unauthorized'}), 403)

        if data.get('type') != 'quiz':
            return _cors_response(jsonify({'ok': False, 'error': 'unknown_type'}), 400)

        # 1. Сохраняем локально
        entry = add_quiz_entry(data)

        # 2. Пересылаем в Google Sheets
        sheets_ok = forward_to_sheets(data)

        # 3. Уведомление в Telegram
        text = fmt_quiz(entry)
        try:
            bot.send_message(ADMIN_ID, text, parse_mode='HTML')
        except Exception as e:
            print(f"⚠️  Telegram: {e}")

        return _cors_response(jsonify({
            'ok': True,
            'type': 'quiz',
            'sheets': sheets_ok
        }))

    except Exception as e:
        print(f"❌ /api/submit error: {e}")
        return _cors_response(jsonify({'ok': False, 'error': str(e)}), 500)

# ─── Проверка статуса заявки ───
@app.route('/api/status', methods=['GET', 'OPTIONS'])
def api_status():
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    phone = request.args.get('phone', '').strip()
    if not phone:
        return _cors_response(jsonify({'status': 'pending'}))

    status = check_status_in_sheets(phone)
    if status is None:
        status = get_status_by_phone(phone)

    if status in ('approved', 'rejected'):
        set_status_by_phone(phone, status)

    return _cors_response(jsonify({'status': status}))

# ─── Отзыв заявки ───
@app.route('/api/revoke', methods=['POST', 'OPTIONS'])
def api_revoke():
    if request.method == 'OPTIONS':
        return _cors_response(jsonify({'ok': True}))

    try:
        data = request.get_json(force=True, silent=True) or {}
        phone = str(data.get('phone', '')).strip()
        name  = str(data.get('name', '—')).strip()

        if not phone:
            return _cors_response(jsonify({'ok': False, 'error': 'no_phone'}), 400)

        deleted = delete_by_phone(phone)

        try:
            bot.send_message(
                ADMIN_ID,
                f"🔄 <b>Заявка ОТОЗВАНА</b>\n\n"
                f"👤 <b>Имя:</b> {name}\n"
                f"📱 <b>Телефон:</b> <code>{phone}</code>\n"
                f"🕐 <b>Время отзыва:</b> {_now_str()}",
                parse_mode='HTML'
            )
        except Exception as e:
            print(f"⚠️  Telegram revoke notify: {e}")

        return _cors_response(jsonify({'ok': True, 'deleted': deleted}))

    except Exception as e:
        return _cors_response(jsonify({'ok': False, 'error': str(e)}), 500)

def _cors_response(response, status=200):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Admin-Token'
    response.status_code = status
    return response

# ════════════════════════════════════════
#  ФОРМАТИРОВАНИЕ УВЕДОМЛЕНИЯ
# ════════════════════════════════════════
LANG_FLAGS = {'ru': '🇷🇺 Русский', 'kz': '🇰🇿 Қазақша'}

def fmt_quiz(d):
    flag = LANG_FLAGS.get(d.get('lang', 'ru'), '🌐')
    utm_str = ''
    try:
        utm = json.loads(d.get('utm', '{}')) if isinstance(d.get('utm'), str) else d.get('utm', {})
        if utm:
            utm_str = f"\n📊 <b>UTM:</b> {json.dumps(utm, ensure_ascii=False)}"
    except Exception:
        pass
    return (
        f"🚨 <b>НОВАЯ ЗАЯВКА С КВИЗА!</b>\n"
        f"{'─' * 30}\n\n"
        f"👤 <b>Имя:</b> {d.get('name', '—')}\n"
        f"📱 <b>Контакт:</b> <code>{d.get('phone', '—')}</code>\n"
        f"🏢 <b>Бизнес:</b> {d.get('business', '—')}\n"
        f"🌆 <b>Город:</b> {d.get('city', '—')}\n"
        f"💰 <b>Бюджет на таргет:</b> {d.get('budget', '—')}\n"
        f"💵 <b>Средний чек:</b> {d.get('avg_check', '—')}\n"
        f"📍 <b>Клиенты сейчас:</b> {d.get('client_source', '—')}\n"
        f"⚠️ <b>Главная проблема:</b> {d.get('main_problem', '—')}\n"
        f"👥 <b>Менеджер:</b> {d.get('has_sales', '—')}\n"
        f"🎯 <b>Цель:</b> {d.get('goal', '—')}\n"
        f"🌐 <b>Язык:</b> {flag}\n"
        f"🕐 <b>Время:</b> {d.get('time', '—')} (Алматы)"
        f"{utm_str}\n\n"
        f"⚡️ <b>Свяжись в течение 5 минут!</b>"
    )

# ════════════════════════════════════════
#  GOOGLE SHEETS — ЧТЕНИЕ ДЛЯ СТАТИСТИКИ
# ════════════════════════════════════════

def fetch_sheet_csv(gid):
    if not SHEET_ID or not gid:
        return None
    url = (f'https://docs.google.com/spreadsheets/d/{SHEET_ID}'
           f'/export?format=csv&gid={gid}')
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as r:
            return r.read().decode('utf-8')
    except Exception:
        return None

def parse_quiz_sheet(csv_text):
    if not csv_text:
        return []
    import csv
    quizzes = []
    lines = csv_text.strip().split('\n')
    for line in lines[1:]:
        try:
            row = next(csv.reader(io.StringIO(line)))
        except StopIteration:
            continue
        cols = [c.strip() for c in row]
        if len(cols) < 3 or not cols[1]:
            continue
        quizzes.append({
            'name':          cols[1]  if len(cols) > 1  else '—',
            'phone':         cols[2]  if len(cols) > 2  else '—',
            'business':      cols[3]  if len(cols) > 3  else '—',
            'city':          cols[4]  if len(cols) > 4  else '—',
            'budget':        cols[5]  if len(cols) > 5  else '—',
            'avg_check':     cols[6]  if len(cols) > 6  else '—',
            'client_source': cols[7]  if len(cols) > 7  else '—',
            'main_problem':  cols[8]  if len(cols) > 8  else '—',
            'has_sales':     cols[9]  if len(cols) > 9  else '—',
            'goal':          cols[10] if len(cols) > 10 else '—',
            'lang':          'kz' if 'Қазақша' in (cols[11] if len(cols) > 11 else '') else 'ru',
            'time':          cols[12] if len(cols) > 12 else '—',
            'status':        cols[13] if len(cols) > 13 else 'pending',
        })
    return quizzes

def load_db_sheets():
    quiz_csv = fetch_sheet_csv(SHEET_QUIZ_GID)
    if quiz_csv is not None:
        return {'quizzes': parse_quiz_sheet(quiz_csv), 'source': '📊 Google Sheets'}
    db = load_db()
    db['source'] = '💾 Локальный файл'
    return db

# ════════════════════════════════════════
#  СТАТИСТИКА
# ════════════════════════════════════════

def build_stats(db, title="ЗА ВСЁ ВРЕМЯ"):
    quizzes = db.get('quizzes', [])
    biz_cnt    = defaultdict(int)
    city_cnt   = defaultdict(int)
    budget_cnt = defaultdict(int)
    lang_cnt   = defaultdict(int)
    for q in quizzes:
        biz_cnt[q.get('business', '—')] += 1
        city_cnt[q.get('city', '—')] += 1
        budget_cnt[q.get('budget', '—')] += 1
        lang_cnt[q.get('lang', 'ru')] += 1

    def top(cnt):
        return '\n'.join(f"  • {k[:30]}: <b>{v}</b>" for k, v in sorted(cnt.items(), key=lambda x: -x[1])) or '  Нет данных'

    return (
        f"📊 <b>СТАТИСТИКА {title}</b>\n{'─'*32}\n\n"
        f"📋 <b>Всего заявок:</b> <b>{len(quizzes)}</b>\n"
        f"🌐 <b>По языку:</b>\n"
        f"  🇷🇺 Русский: {lang_cnt.get('ru', 0)}\n"
        f"  🇰🇿 Казахский: {lang_cnt.get('kz', 0)}\n\n"
        f"🏢 <b>По типу бизнеса:</b>\n{top(biz_cnt)}\n\n"
        f"🌆 <b>По городам:</b>\n{top(city_cnt)}\n\n"
        f"💰 <b>По бюджету:</b>\n{top(budget_cnt)}"
    )

# ════════════════════════════════════════
#  ЭКСПОРТ
# ════════════════════════════════════════

def export_excel(db):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl не установлен"

    wb = Workbook()
    ws = wb.active
    ws.title = "Квиз-заявки"
    hdr_fill = PatternFill("solid", fgColor="7C3AED")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    bd = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )
    row_fills = [PatternFill("solid", fgColor="F5F0FF"), PatternFill("solid", fgColor="FFFFFF")]
    headers = ['#','Имя','Телефон/TG','Бизнес','Город','Бюджет','Средний чек','Клиенты','Проблема','Менеджер','Цель','Язык','Дата','Статус']
    ws.append(headers)
    for i in range(1, len(headers)+1):
        c = ws.cell(1, i)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center'); c.border = bd
    for idx, q in enumerate(db.get('quizzes', []), 1):
        row = [idx, q.get('name',''), q.get('phone',''), q.get('business',''), q.get('city',''),
               q.get('budget',''), q.get('avg_check',''), q.get('client_source',''),
               q.get('main_problem',''), q.get('has_sales',''), q.get('goal',''),
               'Русский' if q.get('lang')=='ru' else 'Қазақша', q.get('time',''), q.get('status','pending')]
        ws.append(row)
        fill = row_fills[idx % 2]
        for i in range(1, len(headers)+1):
            c = ws.cell(idx+1, i); c.fill = fill; c.border = bd
            c.alignment = Alignment(vertical='center', wrap_text=True)
    col_widths = [5,20,22,20,14,22,18,24,28,18,28,12,18,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, None

def export_csv(db):
    lines = ['\ufeff' + 'Имя,Телефон/TG,Бизнес,Город,Бюджет,Средний чек,Клиенты,Проблема,Менеджер,Цель,Язык,Время,Статус']
    def esc(s): return '"' + str(s or '').replace('"','""') + '"'
    for q in db.get('quizzes', []):
        lines.append(','.join([
            esc(q.get('name','')), esc(q.get('phone','')), esc(q.get('business','')),
            esc(q.get('city','')), esc(q.get('budget','')), esc(q.get('avg_check','')),
            esc(q.get('client_source','')), esc(q.get('main_problem','')),
            esc(q.get('has_sales','')), esc(q.get('goal','')),
            esc('Русский' if q.get('lang')=='ru' else 'Қазақша'),
            esc(q.get('time','')), esc(q.get('status','pending')),
        ]))
    return '\n'.join(lines)

# ════════════════════════════════════════
#  TELEGRAM КОМАНДЫ
# ════════════════════════════════════════

def is_admin(msg):
    return str(msg.chat.id) == str(ADMIN_ID)

@bot.message_handler(commands=['start'])
def cmd_start(msg):
    if not is_admin(msg): return
    bot.send_message(msg.chat.id,
        "👋 <b>Dias TargetPro Bot v5.0</b>\n\n"
        "📋 /stats — Статистика за всё время\n"
        "📅 /today — За сегодня\n"
        "📆 /week — За 7 дней\n"
        "📋 /leads — Последние 10 заявок\n"
        "📊 /excel — Скачать Excel\n"
        "📄 /csv — Скачать CSV\n"
        "❓ /help — Помощь",
        parse_mode='HTML'
    )

@bot.message_handler(commands=['stats'])
def cmd_stats(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    bot.send_message(msg.chat.id, build_stats(db) + f"\n\n🗄 {db.get('source','')}", parse_mode='HTML')

@bot.message_handler(commands=['today'])
def cmd_today(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    today = datetime.now().strftime('%d.%m.%Y')
    filtered = {'quizzes': [q for q in db.get('quizzes',[]) if q.get('time','').startswith(today)]}
    if not filtered['quizzes']:
        bot.send_message(msg.chat.id, f"📅 Сегодня ({today}) заявок нет."); return
    bot.send_message(msg.chat.id, build_stats(filtered, f"СЕГОДНЯ ({today})"), parse_mode='HTML')

@bot.message_handler(commands=['week'])
def cmd_week(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    def in_week(q):
        try:
            dt = datetime.strptime(q.get('time','')[:10], '%d.%m.%Y')
            return dt >= datetime.now() - timedelta(days=7)
        except: return False
    filtered = {'quizzes': [q for q in db.get('quizzes',[]) if in_week(q)]}
    if not filtered['quizzes']:
        bot.send_message(msg.chat.id, "📅 За 7 дней заявок нет."); return
    bot.send_message(msg.chat.id, build_stats(filtered, "ЗА 7 ДНЕЙ"), parse_mode='HTML')

@bot.message_handler(commands=['leads'])
def cmd_leads(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    quizzes = db.get('quizzes', [])[-10:]
    if not quizzes:
        bot.send_message(msg.chat.id, "📋 Заявок пока нет."); return
    lines = [f"📋 <b>ПОСЛЕДНИЕ {len(quizzes)} ЗАЯВОК</b>\n"]
    for i, q in enumerate(reversed(quizzes), 1):
        flag = '🇷🇺' if q.get('lang')=='ru' else '🇰🇿'
        status_emoji = {'approved':'✅','rejected':'❌','pending':'🔍'}.get(q.get('status','pending'),'🔍')
        lines.append(f"{i}. {flag} <b>{q['name']}</b> {status_emoji}")
        lines.append(f"   📱 <code>{q['phone']}</code>")
        lines.append(f"   🏢 {q.get('business','—')} • {q.get('city','—')}")
        lines.append(f"   🕐 {q.get('time','')}\n")
    bot.send_message(msg.chat.id, '\n'.join(lines), parse_mode='HTML')

@bot.message_handler(commands=['excel'])
def cmd_excel(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    buf, err = export_excel(db)
    if err:
        bot.send_message(msg.chat.id, f"❌ {err}"); return
    fname = f"dias_quiz_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    bot.send_document(msg.chat.id, (fname, buf),
        caption=f"📊 Excel • Заявок: {len(db.get('quizzes',[]))} • {_now_str()}")

@bot.message_handler(commands=['csv'])
def cmd_csv(msg):
    if not is_admin(msg): return
    db = load_db_sheets()
    buf = io.BytesIO(export_csv(db).encode('utf-8-sig'))
    fname = f"dias_quiz_{datetime.now().strftime('%d%m%Y_%H%M')}.csv"
    bot.send_document(msg.chat.id, (fname, buf), caption=f"📄 CSV • {_now_str()}")

@bot.message_handler(commands=['help'])
def cmd_help(msg):
    if not is_admin(msg): return
    bot.send_message(msg.chat.id,
        "ℹ️ <b>Dias TargetPro Bot v5.0</b>\n\n"
        "<b>Переменные окружения Render:</b>\n"
        "<code>BOT_TOKEN</code> — токен бота\n"
        "<code>ADMIN_ID</code> — Telegram ID\n"
        "<code>SECRET</code> — секрет для /api/submit\n"
        "<code>APPS_SCRIPT_URL</code> — URL Apps Script\n"
        "<code>ADMIN_PASSWORD_HASH</code> — SHA-256 хэш пароля admin\n"
        "<code>WA_NUMBER</code> — номер WhatsApp (опц.)\n"
        "<code>SHEET_ID</code> — ID таблицы (опц.)\n"
        "<code>SHEET_QUIZ_GID</code> — GID листа (опц.)",
        parse_mode='HTML'
    )

@bot.message_handler(func=lambda m: True)
def unknown(msg):
    if is_admin(msg):
        bot.send_message(msg.chat.id, "Используй /help")

# ════════════════════════════════════════
#  ЗАПУСК: Flask + Telegram одновременно
# ════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 60)
    print("  DIAS TARGETPRO v5.0 — Flask API + Telegram Bot")
    print("  🔒 Все секреты загружены из переменных окружения")
    print("=" * 60)
    print(f"  🌐 API сервер: http://0.0.0.0:{PORT}")
    print(f"  🤖 Telegram Admin ID: {ADMIN_ID}")
    print(f"  🔑 Secret: {SECRET[:6]}***")
    print(f"  📊 Apps Script: ✅ настроен")
    print("=" * 60)

    def run_bot():
        print("🤖 Telegram бот запущен...")
        bot.infinity_polling(timeout=60, long_polling_timeout=60)

    bot_thread = threading.Thread(target=run_bot, daemon=True)
    bot_thread.start()

    print(f"🌐 Flask сервер запущен на порту {PORT}...")
    app.run(host='0.0.0.0', port=PORT, debug=False)
